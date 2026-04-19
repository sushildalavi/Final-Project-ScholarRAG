#!/usr/bin/env python3
"""Generate 3 per-coder labeling workbooks from a claim_annotations.csv.

Input:  CSV produced by `extract_public_mode_eval.py` (or any file with the
        columns query_id, query, claim_id, claim_text, evidence_text,
        msa_M, msa_S, msa_A, support_label).

Output: Three .xlsx files — one per coder (A, B, C). Each workbook has:
    Instructions · Rubric · Labeling · Summary

Each coder sees only their own label column. Row order is identical across
all three files so labels can be merged by claim_id afterwards.

Run:
    python backend/scripts/generate_labeling_workbooks.py \
        --claims Evaluation/data/calibration/public_mode/claim_annotations.csv \
        --out-dir Evaluation/data/calibration/public_mode \
        --only-unlabeled
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation


FONT = "Arial"
ILLEGAL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def clean(s: object) -> str:
    if not isinstance(s, str):
        return ""
    s = s.replace("\r", " ").replace("\n", " ").replace("\t", " ")
    return ILLEGAL.sub("", s).strip()


def trunc(s: object, n: int = 900) -> str:
    s = clean(s)
    return s if len(s) <= n else s[:n] + " … [truncated]"


def build_workbook(coder_id: str, claims: pd.DataFrame, out_dir: Path) -> Path:
    wb = Workbook()

    ins = wb.active
    ins.title = "Instructions"
    t = ins["A1"]
    t.value = f"ScholarRAG Labeling — Coder {coder_id}"
    t.font = Font(name=FONT, size=16, bold=True, color="990000")
    ins.merge_cells("A1:D1")

    meta = [
        ("Coder ID:", coder_id),
        ("Total claims:", f"{len(claims)}"),
        ("Your task:", "Label each claim as 'supported' or 'unsupported' based on the cited evidence."),
        ("Time estimate:", f"~30-60 seconds per claim · ~{max(1, round(len(claims) * 0.8 / 60))} hour(s) total."),
        ("Rule 0:", "Do NOT share or compare labels with the other coders until everyone is done."),
    ]
    for i, (k, v) in enumerate(meta, start=3):
        ins.cell(row=i, column=1, value=k).font = Font(name=FONT, size=11, bold=True)
        ins.cell(row=i, column=2, value=v).font = Font(name=FONT, size=11)
        ins.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)

    ins.cell(row=10, column=1, value="Step-by-step").font = Font(name=FONT, size=13, bold=True, color="990000")
    steps = [
        "1.  Read the 'Rubric' tab once before labeling.",
        "2.  Open the 'Labeling' tab. Each row = one claim + one evidence passage.",
        "3.  Read col E (claim_text). Identify the specific factual assertion.",
        "4.  Read col F (evidence_text). Does this passage actually support the claim?",
        "5.  Pick your label via the dropdown in col G (Your Label).",
        "6.  Label ONLY your assigned column. Do not look at other raters before committing.",
        "7.  Write a short note in col H (Notes) if uncertain.",
        "8.  Never type free text in col G — use the dropdown only.",
        "9.  Do not skip rows. If genuinely cannot decide, pick the most likely label and explain in Notes.",
        "10. Save often (Cmd-S / Ctrl-S).",
        "11. When done, save and send the file back. Do not rename it.",
    ]
    for i, s in enumerate(steps, start=12):
        ins.cell(row=i, column=1, value=s).font = Font(name=FONT, size=11)
        ins.merge_cells(start_row=i, start_column=1, end_row=i, end_column=10)

    ins.cell(row=12 + len(steps) + 1, column=1, value="Do / Don't").font = Font(
        name=FONT, size=13, bold=True, color="990000"
    )
    rules = [
        "DO  — Use ONLY the evidence passage in col F. No outside knowledge, no Google.",
        "DO  — Label system abstentions ('I only found X in profile/course context…') as UNSUPPORTED.",
        "DO  — Trust your first read; don't second-guess unless genuinely ambiguous.",
        "DON'T — Look at other coders' files or discuss before everyone finishes.",
        "DON'T — Type anything other than the dropdown options into col G.",
        "DON'T — Skip rows. Every row must have a label before you send the file back.",
    ]
    for i, r in enumerate(rules, start=12 + len(steps) + 3):
        ins.cell(row=i, column=1, value=r).font = Font(name=FONT, size=11)
        ins.merge_cells(start_row=i, start_column=1, end_row=i, end_column=10)

    ins.column_dimensions["A"].width = 22
    for col in "BCDEFGH":
        ins.column_dimensions[col].width = 20

    rub = wb.create_sheet("Rubric")
    rub["A1"].value = "Rubric — Supported vs Unsupported"
    rub["A1"].font = Font(name=FONT, size=16, bold=True, color="990000")
    rub.merge_cells("A1:C1")
    for i, h in enumerate(["Label", "When to use", "Quick test"], start=1):
        c = rub.cell(row=3, column=i, value=h)
        c.font = Font(name=FONT, size=12, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="990000")
        c.alignment = Alignment(horizontal="center", vertical="center")

    defs = [
        (
            "supported",
            "Evidence explicitly states or strongly implies the specific factual content of the claim. A reviewer would accept this passage as the source.",
            "Can you point to a sentence in the evidence that a reviewer would accept as grounding for this claim?",
        ),
        (
            "unsupported",
            "Evidence does NOT substantiate the claim. Includes: (a) different topic/paper; (b) system abstentions; (c) topically related but does not assert the specific claim; (d) evidence contradicts the claim.",
            "If the claim used this evidence as its only source, would a reviewer call it hallucinated?",
        ),
    ]
    for i, (lbl, when, test) in enumerate(defs, start=4):
        rub.cell(row=i, column=1, value=lbl).font = Font(name=FONT, size=11, bold=True)
        rub.cell(row=i, column=2, value=when).font = Font(name=FONT, size=11)
        rub.cell(row=i, column=3, value=test).font = Font(name=FONT, size=11, italic=True)
        rub.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        rub.cell(row=i, column=3).alignment = Alignment(wrap_text=True, vertical="top")
        rub.row_dimensions[i].height = 120

    rub.cell(row=7, column=1, value="Worked examples").font = Font(
        name=FONT, size=13, bold=True, color="990000"
    )
    examples = [
        ("Claim", "Evidence snippet", "Correct label", "Why"),
        (
            "BERT is pre-trained using a masked language model objective.",
            "We introduce BERT ... pre-trained using a masked language model (MLM) objective.",
            "supported",
            "Evidence explicitly states the MLM objective.",
        ),
        (
            "Self-attention computes weighted sums of Value vectors.",
            "The paper evaluates generation on summarization tasks using ROUGE.",
            "unsupported",
            "Evidence is on a different topic.",
        ),
        (
            "I only found Main Idea mentioned in profile/course context.",
            "Any evidence snippet.",
            "unsupported",
            "System-level abstention = never supported.",
        ),
        (
            "The paper achieves SOTA on GLUE.",
            "We evaluate on GLUE; our model performs competitively across most tasks.",
            "unsupported",
            "Evidence says 'competitively', not SOTA — claim overstates.",
        ),
    ]
    for i, row in enumerate(examples):
        for j, cell in enumerate(row):
            c = rub.cell(row=9 + i, column=j + 1, value=cell)
            c.font = Font(name=FONT, size=10, bold=(i == 0))
            c.alignment = Alignment(wrap_text=True, vertical="top")
            if i == 0:
                c.fill = PatternFill("solid", start_color="E8E8E8")
        rub.row_dimensions[9 + i].height = 55 if i > 0 else 20
    rub.column_dimensions["A"].width = 32
    rub.column_dimensions["B"].width = 48
    rub.column_dimensions["C"].width = 16
    rub.column_dimensions["D"].width = 36

    lab = wb.create_sheet("Labeling")
    headers = ["#", "query_id", "query", "claim_id", "claim_text", "evidence_text",
               f"Your Label (Coder {coder_id})", "Notes"]
    for i, h in enumerate(headers, start=1):
        c = lab.cell(row=1, column=i, value=h)
        c.font = Font(name=FONT, size=11, bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="990000")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    for i, row in claims.iterrows():
        r = i + 2
        lab.cell(row=r, column=1, value=int(i + 1))
        lab.cell(row=r, column=2, value=clean(row.get("query_id")))
        lab.cell(row=r, column=3, value=trunc(row.get("query"), 200))
        lab.cell(row=r, column=4, value=clean(row.get("claim_id")))
        lab.cell(row=r, column=5, value=trunc(row.get("claim_text"), 600))
        lab.cell(row=r, column=6, value=trunc(row.get("evidence_text"), 900))
        for c_i in range(1, 9):
            cell = lab.cell(row=r, column=c_i)
            cell.font = Font(name=FONT, size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin

    last_row = len(claims) + 1
    dv = DataValidation(type="list", formula1='"supported,unsupported"', allow_blank=True)
    dv.error = "Pick supported or unsupported from the dropdown."
    dv.errorTitle = "Invalid label"
    dv.add(f"G2:G{last_row}")
    lab.add_data_validation(dv)

    green_fill = PatternFill("solid", start_color="C6EFCE")
    red_fill = PatternFill("solid", start_color="FFC7CE")
    lab.conditional_formatting.add(f"G2:G{last_row}", CellIsRule(operator="equal", formula=['"supported"'], fill=green_fill))
    lab.conditional_formatting.add(f"G2:G{last_row}", CellIsRule(operator="equal", formula=['"unsupported"'], fill=red_fill))

    widths = {"A": 5, "B": 8, "C": 35, "D": 10, "E": 55, "F": 80, "G": 18, "H": 35}
    for col, w in widths.items():
        lab.column_dimensions[col].width = w
    lab.freeze_panes = "G2"
    lab.row_dimensions[1].height = 30

    smr = wb.create_sheet("Summary")
    smr["A1"].value = f"Your progress (Coder {coder_id})"
    smr["A1"].font = Font(name=FONT, size=16, bold=True, color="990000")
    smr.merge_cells("A1:C1")
    stats = [
        ("Total claims", f"=COUNTA(Labeling!A2:A{last_row})", "int"),
        ("Labeled so far", f"=COUNTA(Labeling!G2:G{last_row})", "int"),
        ("Remaining", f"=COUNTA(Labeling!A2:A{last_row})-COUNTA(Labeling!G2:G{last_row})", "int"),
        ("% complete",
         f"=IFERROR(COUNTA(Labeling!G2:G{last_row})/COUNTA(Labeling!A2:A{last_row}),0)", "pct"),
        ("", "", ""),
        ("You labeled: supported", f'=COUNTIF(Labeling!G2:G{last_row},"supported")', "int"),
        ("You labeled: unsupported", f'=COUNTIF(Labeling!G2:G{last_row},"unsupported")', "int"),
        ("Your supported %",
         f'=IFERROR(COUNTIF(Labeling!G2:G{last_row},"supported")/COUNTA(Labeling!G2:G{last_row}),0)', "pct"),
    ]
    for i, (k, v, fmt) in enumerate(stats, start=3):
        smr.cell(row=i, column=1, value=k).font = Font(name=FONT, size=11, bold=True)
        c = smr.cell(row=i, column=2, value=v)
        c.font = Font(name=FONT, size=11)
        if fmt == "pct":
            c.number_format = "0.0%"
        elif fmt == "int":
            c.number_format = "0"
    smr.column_dimensions["A"].width = 32
    smr.column_dimensions["B"].width = 16

    out_path = out_dir / f"ScholarRAG_PublicMode_Coder_{coder_id}.xlsx"
    wb.save(out_path)
    return out_path


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--claims", required=True, help="Path to claim_annotations.csv")
    parser.add_argument("--out-dir", required=True, help="Directory to write the three workbook files")
    parser.add_argument(
        "--only-unlabeled",
        action="store_true",
        help="Only include rows where support_label is empty/NaN",
    )
    args = parser.parse_args()

    df = pd.read_csv(args.claims)
    if args.only_unlabeled:
        df = df[~df["support_label"].isin(["supported", "unsupported"])].copy()
    df = df.dropna(subset=["claim_text", "evidence_text"]).reset_index(drop=True)

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    print(f"Generating workbooks for {len(df)} claims -> {out_dir}")
    for coder in ("A", "B", "C"):
        path = build_workbook(coder, df, out_dir)
        print(f"  wrote {path}")


if __name__ == "__main__":
    main()
